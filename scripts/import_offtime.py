"""Import off-time (home) days + holidays from the team's reference Gantt
workbook into the app.

Re-run this any time the source workbook changes — it fully replaces the
previously-imported holidays list and every Sambatz person's off-time
entries with what's currently in the file, so it's safe to run repeatedly
as the schedule gets revised.

Usage:
    python scripts/import_offtime.py "<path to .xlsx>" [--apply-local] [--sql-only]

What it reads (sheet "מכלול כללי חדש" of the workbook):
  - Row 2: date header (one column per day)
  - Row 5: holiday name per date, where present
  - Column A: role, column B: full name, one row per person
  - A cell with 'V' in a person's row/date column means that person is
    AT BASE that day — anything else (blank, 'X', ...) means AT HOME
    (off duty). Only an explicit V counts as base; the default is home.
  - Every day strictly before the workbook's earliest date column is
    treated as pre-term leave for every matched person — the reserve
    term hasn't started yet, so there's no "at base" to default to.

What it writes (into this scripts/ dir, always):
  - seed_offtime.sql   — replaces public.time_off for every person found
                          in the workbook (delete-then-insert, scoped to
                          just those people) and prints to stdout too
  - holiday dates       — printed as a ready-to-paste TS object literal
                          for src/lib/holidays.ts (that file isn't
                          auto-written since it's committed source, not
                          data — copy the block over by hand)

Person matching is by exact full_name against whatever's currently in the
local Supabase `profiles` table (via `supabase db query --local`), so a
name that doesn't exist yet in the app is skipped with a warning rather
than silently failing.
"""

import argparse
import subprocess
import sys
from datetime import timedelta
from pathlib import Path

import openpyxl

# Safely before any realistic "today" the app will be viewed on, so the
# pre-term leave range always covers "now" regardless of when this script
# actually gets run relative to the term start date.
PRE_TERM_ANCHOR_ISO = "2000-01-01"

SHEET_NAME = "מכלול כללי חדש"
DATE_ROW = 2
DATE_ROW_LABEL = "תאריך"
HOLIDAY_ROW = 5
FIRST_DATA_ROW = 6
# Only an explicit 'V' means at-base — blank, 'X', or anything else means
# at home. This inverts the earlier "blank defaults to base" assumption per
# explicit correction: not-X-and-not-V is the same as X (at home).
BASE_MARK = "V"

DOCKER_EXE = r"C:\Users\Barash\AppData\Local\Programs\DockerDesktop\resources\bin\docker.exe"
DB_CONTAINER = "supabase_db_shift-manager"


def run_psql(sql: str) -> str:
    """Run SQL against the local Supabase DB via `docker exec ... psql`."""
    result = subprocess.run(
        [DOCKER_EXE, "exec", "-i", DB_CONTAINER, "psql", "-U", "postgres", "-d", "postgres", "-t", "-A", "-c", sql],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr)
    return result.stdout


def apply_sql_file(path: Path) -> None:
    with open(path, "r", encoding="utf-8") as f:
        result = subprocess.run(
            [DOCKER_EXE, "exec", "-i", DB_CONTAINER, "psql", "-U", "postgres", "-d", "postgres"],
            stdin=f,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
    print(result.stdout)
    if result.returncode != 0:
        print(result.stderr, file=sys.stderr)
        sys.exit(result.returncode)


def load_known_profile_names() -> set[str]:
    """Full names currently in the local profiles table."""
    try:
        out = run_psql("select full_name from public.profiles;")
    except Exception as e:
        print(f"warning: could not query local profiles ({e}), skipping name validation", file=sys.stderr)
        return set()
    return {line.strip() for line in out.splitlines() if line.strip()}


def extract(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]

    max_col = ws.max_column

    # The sheet layout has shifted between file variants (extra summary
    # columns prepended in some) — locate the "תאריך" label in the date row
    # rather than assuming a fixed column, and take the name column as the
    # one right before it, dates starting right after.
    name_col = None
    for c in range(1, max_col + 1):
        if ws.cell(row=DATE_ROW, column=c).value == DATE_ROW_LABEL:
            name_col = c
            break
    if name_col is None:
        raise ValueError(f"couldn't find the '{DATE_ROW_LABEL}' label in row {DATE_ROW} of '{SHEET_NAME}'")
    first_date_col = name_col + 1

    dates = {}
    for c in range(first_date_col, max_col + 1):
        d = ws.cell(row=DATE_ROW, column=c).value
        if hasattr(d, "date"):
            dates[c] = d.date()

    holidays = {}
    for c, d in dates.items():
        h = ws.cell(row=HOLIDAY_ROW, column=c).value
        if h:
            holidays[d.isoformat()] = h

    people = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        if not name:
            continue
        home_dates = []
        for c, d in dates.items():
            v = ws.cell(row=r, column=c).value
            if v != BASE_MARK:
                home_dates.append(d)
        home_dates.sort()
        ranges = []
        for d in home_dates:
            if ranges and (d - ranges[-1][1]).days == 1:
                ranges[-1] = (ranges[-1][0], d)
            else:
                ranges.append((d, d))
        people[name] = ranges

    term_start = min(dates.values()) if dates else None
    return holidays, people, term_start


def build_sql(people: dict[str, list], known_names: set[str], term_start) -> str:
    matched = {name: ranges for name, ranges in people.items() if not known_names or name in known_names}

    if term_start is not None:
        from datetime import date as _date

        anchor = _date.fromisoformat(PRE_TERM_ANCHOR_ISO)
        pre_term_end = term_start - timedelta(days=1)
        if pre_term_end >= anchor:
            for name in matched:
                matched[name] = [(anchor, pre_term_end)] + matched[name]
    skipped = [name for name in people if known_names and name not in known_names]
    for name in skipped:
        print(f"warning: '{name}' not found in profiles table, skipping", file=sys.stderr)

    names_sql = ", ".join(f"'{n.replace(chr(39), chr(39) * 2)}'" for n in matched)
    delete_sql = (
        f"delete from public.time_off using public.profiles p "
        f"where p.id = time_off.user_id and p.full_name in ({names_sql});\n\n"
        if matched
        else ""
    )

    rows = []
    for name, ranges in matched.items():
        esc = name.replace("'", "''")
        for s, e in ranges:
            rows.append(f"  ('{esc}', date '{s.isoformat()}', date '{e.isoformat()}')")

    if not rows:
        return delete_sql + "-- no off-time ranges found to insert\n"

    insert_sql = (
        "insert into public.time_off (user_id, start_date, end_date)\n"
        "select p.id, v.start_date, v.end_date from (values\n"
        + ",\n".join(rows)
        + "\n) as v(full_name, start_date, end_date)\n"
        "join public.profiles p on p.full_name = v.full_name;\n"
    )
    return delete_sql + insert_sql


def build_holidays_ts(holidays: dict[str, str]) -> str:
    lines = ["export const ISRAELI_HOLIDAYS: Record<string, string> = {"]
    for date in sorted(holidays):
        name = holidays[date].replace('"', '\\"')
        lines.append(f'  "{date}": "{name}",')
    lines.append("};")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--apply-local", action="store_true", help="also run the SQL against the local DB")
    args = parser.parse_args()

    holidays, people, term_start = extract(args.xlsx)
    known_names = load_known_profile_names()
    sql = build_sql(people, known_names, term_start)

    out_dir = Path(__file__).resolve().parent
    sql_path = out_dir / "seed_offtime.sql"
    sql_path.write_text(sql, encoding="utf-8")

    print("=" * 60)
    print(f"Holidays found: {len(holidays)}")
    print("Paste into src/lib/holidays.ts:")
    print()
    print(build_holidays_ts(holidays))
    print()
    print("=" * 60)
    print(f"SQL written to {sql_path}")
    print(sql)

    if args.apply_local:
        print("Applying to local DB...")
        apply_sql_file(sql_path)


if __name__ == "__main__":
    main()
